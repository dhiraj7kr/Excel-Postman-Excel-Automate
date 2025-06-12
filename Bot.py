import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill

# Load Excel with conversation IDs
df = pd.read_excel('conversation_ids.xlsx')

# Add required columns
df['User'] = ''
df['Bot'] = ''
df['currentUrlLink'] = ''

# API info
url = "https://prod-133.westus.logic.azure.com:443/workflows/1bd55600516f4dc9912ea64a6430f017/triggers/manual/paths/invoke"
params = {
    'api-version': '2016-06-01',
    'sp': '/triggers/manual/run',
    'sv': '1.0',
    'sig': '4OybBdWcrGKHK6gpfEmd96hxrE3pvKjfyAAEZu-8ny8'
}
headers = {
    'Content-Type': 'application/json'
}

# Process each conversation ID
for idx, row in df.iterrows():
    conversation_id = row['cr7de_conversationid']
    payload = {'conversationid': conversation_id}

    try:
        response = requests.post(url, headers=headers, params=params, json=payload)
        data = response.json()

        found = False
        last_user = ''
        last_bot = ''
        last_url = ''

        for msg in data:
            if 'User' in msg and msg['User'].strip():
                last_user = msg['User']
            if 'Currenturl' in msg and msg['Currenturl'].strip():
                last_url = msg['Currenturl']
            if 'Bot' in msg and msg['Bot'].strip():
                last_bot = msg['Bot']

            if 'Bot' in msg and msg['Bot'].strip() == "":
                df.at[idx, 'User'] = last_user
                df.at[idx, 'Bot'] = last_bot
                df.at[idx, 'currentUrlLink'] = last_url

                print(f"Conversation ID: {conversation_id}")
                print(f"User: {last_user}")
                print(f"Bot: {last_bot}")
                print(f"URL: {last_url}")
                print('-' * 60)

                found = True
                break

        if not found:
            print(f"Conversation ID: {conversation_id} - No empty Bot response found.\n{'-' * 60}")

    except Exception as e:
        print(f"Error processing conversation ID {conversation_id}: {e}")
        continue

# Finalize columns
df_final = df[['cr7de_conversationid', 'User', 'Bot', 'currentUrlLink']]
df_final.rename(columns={'cr7de_conversationid': 'ConversationId'}, inplace=True)

# Save to Excel
output_path = 'conversation_ids_output.xlsx'
df_final.to_excel(output_path, index=False)

# Apply formatting using openpyxl
wb = load_workbook(output_path)
ws = wb.active

# Set column widths
column_widths = {
    'A': 25,  # ConversationId
    'B': 40,  # User
    'C': 40,  # Bot
    'D': 50   # currentUrlLink
}

for col_letter, width in column_widths.items():
    ws.column_dimensions[col_letter].width = width

# Set header fill color (light blue)
header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

for cell in ws[1]:  # First row = headers
    cell.fill = header_fill

wb.save(output_path)
print(f"\n✅ Excel saved with blue-colored headers and optimized column widths at: {output_path}")
