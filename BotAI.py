import pandas as pd
import requests
import json
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Category mapping (unchanged)
category_map = {
    "https://www.hollandamerica.com/en/us/booking/summary/Cruise-Summary-Overview/cruise-summary-guestinfo/cruise-summary-options/payment/confirmation": "Confirmation",
    "https://www.hollandamerica.com/en/us/booking/summary/Cruise-Summary-Overview/cruise-summary-guestinfo/cruise-summary-options/payment": "Payment",
    "https://www.hollandamerica.com/en/us/booking/summary/Cruise-Summary-Overview/cruise-summary-guestinfo/cruise-summary-options": "Cruise Protection",
    "https://www.hollandamerica.com/en/us/booking/summary/Cruise-Summary-Overview/cruise-summary-guestinfo": "Guest Info",
    "https://www.hollandamerica.com/en/us/booking/summary/Cruise-Summary-Overview": "Summary",
    "https://www.hollandamerica.com/en/us/booking/choose-your-guests/choose-your-fare/upgrade/we-choose-you-choose/section/deck/room": "Stateroom Number",
    "https://www.hollandamerica.com/en/us/booking/choose-your-guests/choose-your-fare/upgrade/we-choose-you-choose/section/deck": "Stateroom Deck",
    "https://www.hollandamerica.com/en/us/booking/choose-your-guests/choose-your-fare/upgrade/we-choose-you-choose/section": "Stateroom Section",
    "https://www.hollandamerica.com/en/us/booking/choose-your-guests/choose-your-fare/upgrade/we-choose-you-choose": "Stateroom Location",
    "https://www.hollandamerica.com/en/us/booking/choose-your-guests/choose-your-fare": "Fare",
    "https://www.hollandamerica.com/en/us/booking/choose-your-guests": "Guest and Room",
    "https://www.hollandamerica.com/en/us/find-a-cruise/": "Cruise Itinerary",
    "https://www.hollandamerica.com/en/us/find-a-cruise?/": "Book Cruise"
}

def determine_category(link: str) -> str:
    if not isinstance(link, str) or not link.strip():
        return "Knowledge"
    for url, category in category_map.items():
        if link.startswith(url):
            return category
    return "Knowledge"

# Function to summarize conversation using OpenAI
def summarize_with_openai(convo: list) -> str:
    openai_url = "https://api.openai.com/v1/chat/completions"
    openai_key = "****************************************"

    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {openai_key}"
    }

    prompt = (
        "You are given a JSON array representing a full conversation between a user and a bot.\n"
        "If the message contains an empty 'Bot' value, it's considered negative feedback.\n\n"
        "Understand the full conversation and summarize the user's intent and response it is providing and what is the reason for negative feedback.\n"
        "Return a brief summary of what the user was trying to do, and if it's negative feedback, explain what issue they faced and what should be the proper answer.\n\n"
        "Respond ONLY in this format:\n"
        "Summary: <summary text>\n"
        "Satisfaction: Satisfied | Not Satisfied\n"
        "Reason: <reason why not satisfied and about what not satisfied.>\n"
    )

    messages = [
        {"role": "system", "content": prompt},
        {"role": "user", "content": json.dumps(convo, indent=2)}
    ]

    payload = {
        "model": "gpt-4o-mini",
        "messages": messages
    }

    try:
        response = requests.post(openai_url, headers=headers, json=payload)
        result = response.json()
        return result['choices'][0]['message']['content'].strip()
    except Exception as e:
        print(f"OpenAI error: {e}")
        return "Summary: Error summarizing.\nSatisfaction: Unknown"

# Load Excel
df = pd.read_excel('conversation_ids.xlsx')

# Add required columns
df['User'] = ''
df['Bot'] = ''
df['currentUrlLink'] = ''
df['FullConversation'] = ''
df['Category'] = ''
df['SummaryAndSatisfaction'] = ''

# API Info
url = "https://prod-133.westus.logic.azure.com:443/workflows/1bd55600516f4dc9912ea64a6430f017/triggers/manual/paths/invoke"
params = {
    'api-version': '2016-06-01',
    'sp': '/triggers/manual/run',
    'sv': '1.0',
    'sig': '4OybBdWcrGKHK6gpfEmd96hxrE3pvKjfyAAEZu-8ny8'
}
headers = {'Content-Type': 'application/json'}

# Process each conversation ID
for idx, row in df.iterrows():
    conversation_id = row['cr7de_conversationid']
    payload = {'conversationid': conversation_id}

    try:
        response = requests.post(url, headers=headers, params=params, json=payload)
        data = response.json()

        found = False
        last_user = ''
        last_bot = ''
        last_url = ''

        for msg in data:
            if 'User' in msg and msg['User'].strip():
                last_user = msg['User'].strip()
            if 'Bot' in msg and msg['Bot'].strip():
                last_bot = msg['Bot'].strip()
            if 'Currenturl' in msg and msg['Currenturl'].strip():
                last_url = msg['Currenturl'].strip()

            if 'Bot' in msg and msg['Bot'].strip() == "":
                df.at[idx, 'User'] = last_user
                df.at[idx, 'Bot'] = last_bot
                df.at[idx, 'currentUrlLink'] = last_url
                found = True
                break

        # Always store full conversation + category
        full_convo_json = json.dumps(data, indent=4)
        df.at[idx, 'FullConversation'] = full_convo_json
        df.at[idx, 'Category'] = determine_category(last_url)

        # Get summary/satisfaction from OpenAI
        summary_text = summarize_with_openai(data)
        df.at[idx, 'SummaryAndSatisfaction'] = summary_text

        # Logs
        print(f"✅ Conversation ID: {conversation_id}")
        print(summary_text)
        print('-' * 60)

        if not found:
            print(f"⚠️ Conversation ID: {conversation_id} - No empty Bot response found.")

    except Exception as e:
        print(f"❌ Error processing conversation ID {conversation_id}: {e}")
        continue

# Final output
df_final = df[['cr7de_conversationid', 'User', 'Bot', 'currentUrlLink', 'Category', 'SummaryAndSatisfaction', 'FullConversation']]
df_final.rename(columns={'cr7de_conversationid': 'ConversationId'}, inplace=True)

# Save Excel
output_path = 'conversation_ids_output.xlsx'
df_final.to_excel(output_path, index=False)

# Formatting
wb = load_workbook(output_path)
ws = wb.active

column_widths = {
    'A': 25,  # ConversationId
    'B': 40,  # User
    'C': 40,  # Bot
    'D': 50,  # currentUrlLink
    'E': 25,  # Category
    'F': 60,  # SummaryAndSatisfaction
    'G': 120  # FullConversation
}

for col_letter, width in column_widths.items():
    ws.column_dimensions[col_letter].width = width

header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
for cell in ws[1]:
    cell.fill = header_fill

wb.save(output_path)
print(f"\n📑 Excel saved with summary and satisfaction analysis at: {output_path}")
